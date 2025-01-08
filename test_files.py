from zipfile import ZipFile
from openpyxl import load_workbook
from pypdf import PdfReader
import csv
import os


def assert_path():
    zip_path = os.path.abspath("res/files.zip")
    return zip_path


def test_file_pdf(zip_files):
    with ZipFile('res/files.zip', 'r') as z:
        with z.open('morr.pdf') as pdf_file:
            reader = PdfReader(pdf_file)
            page = reader.pages[0]
            assert 'The purpose and significance of this paper' in page.extract_text()


def test_file_csv(zip_files):
    with ZipFile('res/files.zip', 'r') as z:
        with z.open('csv_tab.csv') as csv_file:
            csv_content = csv_file.read().decode('utf-8')
            csv_reader = csv.reader(csv_content.splitlines())
            text = []
            csv_out = []
            for row in csv_reader:
                text.append(row)
                csv_out = ', '.join([', '.join(row) for row in text])
            assert "Павлова" in csv_out


def test_file_xlsx(zip_files):
    with ZipFile('res/files.zip', 'r') as z:
        with z.open('tab.xlsx') as excel_file:
            workbook = load_workbook(excel_file)
            sheet = workbook.active
            assert "Дубина" in sheet.cell(row=9, column=1).value
